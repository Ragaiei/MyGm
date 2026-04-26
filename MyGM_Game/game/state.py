from openpyxl import load_workbook

FILE = "data/wrestlers.xlsx"

def load_wrestlers():
    wb = load_workbook(FILE)
    ws = wb["Wrestlers"]
    data = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0]:
            data.append({
                "name": r[0],
                "power": r[1],
                "pop": r[2],
                "salary": r[4]
            })
    return data

def new_game():
    return {
        "week": 1,
        "money": 500000,
        "pool": load_wrestlers(),
        "brands": {
            "RAW": [],
            "SmackDown": [],
            "ECW": []
        },
        "shows": {
            "RAW": [],
            "SmackDown": [],
            "ECW": []
        },
        "champions": {
            "World": None,
            "IC": None,
            "Tag": None
        },
        "injuries": {},
        "rivalries": {},
        "turn_order": ["ECW","RAW","SmackDown"],
        "turn_index": 0
    }